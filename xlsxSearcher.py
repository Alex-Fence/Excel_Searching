import os
import openpyxl

def find_xlsx_by_sheet_name_and_content(directory, search_term):
    """
    Ищет файлы .xlsx, содержащие заданную строку в имени листа ИЛИ в содержимом листа (без учёта регистра).

    Args:
        directory: Путь к каталогу для поиска.
        search_term: Строка для поиска в именах листов и содержимом.

    Returns:
        Список кортежей (путь_к_файлу, имя_листа), где найдены совпадения.
    """

    found_files = []
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith(".xlsx"):
                filepath = os.path.join(root, file)
                try:
                    workbook = openpyxl.load_workbook(filepath, read_only=True)
                    for sheet_title in workbook.sheetnames:
                        if sheet_title.lower().find(search_term.lower()) != -1: # Проверка имени листа
                            found_files.append((filepath, sheet_title))
                            continue # Переходим к следующему файлу, если найдено в имени листа

                        worksheet = workbook[sheet_title]
                        for row in worksheet.iter_rows():
                            for cell in row:
                                if cell.value and str(cell.value).lower().find(search_term.lower()) != -1:
                                    found_files.append((filepath, sheet_title))
                                    break # Выходим из цикла по строкам после нахождения совпадения
                            if (filepath, sheet_title) in found_files:
                                break # Выходим из цикла по ячейкам после нахождения совпадения
                    workbook.close()
                except Exception as e:
                    print(f"Ошибка при обработке файла {filepath}: {e}")
    return found_files

def write_to_file(file_path, write_str):
    with open(file_path, 'a', encoding='utf-8') as f:
        f.write(write_str)

if __name__ == "__main__":
    search_directory = input("Введите путь к каталогу для поиска: ")
    search_string = input("Введите строку для поиска в именах листов и содержимом (без учёта регистра): ")

    results = find_xlsx_by_sheet_name_and_content(search_directory, search_string)

    if results:
        print("\nНайдены следующие файлы и листы:")
        for file, sheet in results:
            write_to_string = f"Файл: {file}, Лист: {sheet}"+'\n'
            print(write_to_string)
            write_to_file("result.txt", write_to_string)
    else:
        print("\nСовпадений не найдено.")
